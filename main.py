import asyncio
import aiohttp
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from typing import Dict, Any
from telegram import Bot
import gc

class SolanaTokenPriceFetcher:
    def __init__(self, telegram_token: str, telegram_chat_id: str):
        self.rpc_url = "your RPC"
        self.new_price_url = "Your price fetch url eg.dexapi,raydium api, etc" #you can change the price fetch function according to the api documentation.
        self.quote_address = "address of the coin you will be comparing with, eg SOL, USDC, USDT."
        self.file_name = "tokens.xlsx"
        self.telegram_token = telegram_token
        self.telegram_chat_id = telegram_chat_id
        self.bot = Bot(token=telegram_token)

    def initialize_xlsx(self):
        """Create a new tokens.xlsx file with the appropriate columns if it doesn't exist."""
        if not os.path.exists(self.file_name):
            wb = Workbook()
            ws = wb.active
            ws.title = "Token Details"
            columns = ["Token Name", "Symbol", "Mint Address", "Balance", "Price USD", "Market Cap", "Total Value"]
            for idx, col_name in enumerate(columns, start=1):
                ws[f"{get_column_letter(idx)}1"] = col_name
            wb.save(self.file_name)

    def update_xlsx(self, wallet_tokens: Dict[str, float], token_details: list):
        """Update the tokens.xlsx file with new token details and detect significant changes."""
        wb = load_workbook(self.file_name)
        ws = wb.active

        # Read existing data into a dictionary
        existing_data = {
            ws[f"C{row}"].value: {
                "row": row,
                "balance": ws[f"D{row}"].value,
                "total_value": ws[f"G{row}"].value,
            }
            for row in range(2, ws.max_row + 1)
        }

        significant_changes = []

        for details, (mint, balance) in zip(token_details, wallet_tokens.items()):
            if 'error' not in details:
                price_usd = details['priceUsd']
                market_cap = details['marketCap']
                total_value = balance * price_usd

                if mint in existing_data:
                    # Update existing row if values changed
                    row = existing_data[mint]["row"]
                    if (
                        ws[f"D{row}"].value != balance
                        or ws[f"E{row}"].value != price_usd
                        or ws[f"G{row}"].value != total_value
                    ):
                        ws[f"D{row}"].value = balance
                        ws[f"E{row}"].value = price_usd
                        ws[f"F{row}"].value = market_cap
                        ws[f"G{row}"].value = total_value

                    # Check for significant change
                    if existing_data[mint]["total_value"]:
                        previous_value = existing_data[mint]["total_value"]
                        if total_value > 1.5 * previous_value or total_value > 10:
                            significant_changes.append({
                                "name": details["name"],
                                "symbol": details["symbol"],
                                "priceUsd": price_usd,
                                "mint": mint,  # Include mint here
                                "new_total_value": total_value,
                            })
                else:
                    # Add new token row
                    new_row = ws.max_row + 1
                    ws[f"A{new_row}"].value = details["name"]
                    ws[f"B{new_row}"].value = details["symbol"]
                    ws[f"C{new_row}"].value = mint
                    ws[f"D{new_row}"].value = balance
                    ws[f"E{new_row}"].value = price_usd
                    ws[f"F{new_row}"].value = market_cap
                    ws[f"G{new_row}"].value = total_value

        wb.save(self.file_name)
        return significant_changes

    async def get_wallet_tokens(self, wallet_address: str) -> Dict[str, float]:
        try:
            payload = {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "getTokenAccountsByOwner",
                "params": [
                    wallet_address,
                    {"programId": "TokenkegQfeZyiNwAJbNbGKPFXCWuBvf9Ss623VQ5DA"},
                    {"encoding": "jsonParsed", "commitment": "confirmed"},
                ],
            }

            async with aiohttp.ClientSession() as session:
                async with session.post(self.rpc_url, json=payload) as response:
                    data = await response.json()

            token_balances = {}

            if "result" in data and "value" in data["result"]:
                for account in data["result"]["value"]:
                    parsed_info = account["account"]["data"]["parsed"]["info"]
                    mint_address = parsed_info["mint"]
                    token_amount = parsed_info["tokenAmount"]

                    balance = int(token_amount["amount"])
                    decimals = token_amount["decimals"]
                    readable_balance = balance / (10**decimals)

                    if readable_balance > 0:
                        token_balances[mint_address] = readable_balance

            return token_balances

        except Exception as e:
            print(f"Error fetching wallet tokens: {e}")
            return {}

    async def get_token_details(self, session: aiohttp.ClientSession, token_address: str) -> Dict[str, Any]:
        try:
            url = f"{self.new_price_url}/{token_address}?quote_address={self.quote_address}"
            async with session.get(url) as response:
                token_data = await response.json()

            if "priceUsd" in token_data:
                return {
                    "name": token_data.get("name", "Unknown"),
                    "symbol": token_data.get("symbol", "N/A"),
                    "priceUsd": token_data.get("priceUsd", 0),
                    "marketCap": token_data.get("mcap", 0),
                }
            else:
                return {"error": f"Details not found for {token_address}"}

        except Exception as e:
            print(f"Error fetching token details for {token_address}: {e}")
            return {"error": str(e)}

    async def fetch_all_token_details(self, tokens: Dict[str, float]):
        async with aiohttp.ClientSession() as session:
            tasks = [self.get_token_details(session, token) for token in tokens.keys()]
            return await asyncio.gather(*tasks)

    async def send_telegram_message(self, message: str):
        """Send a message to the Telegram bot."""
        await self.bot.send_message(chat_id=self.telegram_chat_id, text=message)

async def main_loop():
    wallet = "your wallet address"
    telegram_token = "your telegram token"
    telegram_chat_id = "your telegram chat id"

    fetcher = SolanaTokenPriceFetcher(telegram_token, telegram_chat_id)

    # Initialize the xlsx file if not present
    fetcher.initialize_xlsx()
    fabulus=0
    while True:
        try:
            wallet_tokens = await fetcher.get_wallet_tokens(wallet)
            if wallet_tokens:
                token_details = await fetcher.fetch_all_token_details(wallet_tokens)
                significant_changes = fetcher.update_xlsx(wallet_tokens, token_details)

                if significant_changes:
                    for token in significant_changes:
                        if token['mint']!='EPjFWdd5AufqSSqeM2qN1xzybapC8G4wEGGkZwyTDt1v':# token you dont want to monitor eg USDC , USDT , currently it is only set to usdc.
                            message = (
                                f"Token: {token['name']} ({token['symbol']})\n"
                                f"Mint Address: {token['mint']}\n"
                                f"New Total Value: ${token['new_total_value']:,.2f}"
                            )
                            await fetcher.send_telegram_message(message)
                            print(message)
            else:
                print("No tokens found in the wallet.")
            await asyncio.sleep(120)
        except Exception as e:
            print(f"Error during execution: {e}")
            await asyncio.sleep(120)
        finally:
            gc.collect()
            print(fabulus)
            fabulus+=1


if __name__ == "__main__":
    asyncio.run(main_loop())
